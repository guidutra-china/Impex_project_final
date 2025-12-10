#!/usr/bin/env python3
"""
Generate Product Import Template Excel File
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Product Import Template"

# Define headers with descriptions
headers = [
    ("A", "Product Name*", "Required. Name of the product"),
    ("B", "Product Code/SKU", "Optional. Will be auto-generated if empty"),
    ("C", "Description", "Optional. Product description"),
    ("D", "Status", "Optional. 'active' or 'inactive' (default: active)"),
    ("E", "Price", "Optional. Price in currency units (e.g., 10.50)"),
    ("F", "Currency Code", "Optional. USD, CNY, EUR, etc. (default: USD)"),
    ("G", "Photo URL", "Optional. URL to download product image"),
    ("H", "Photo Embedded", "Optional. Paste image directly in this cell"),
    ("I", "Supplier Name", "Optional. Must match existing supplier"),
    ("J", "Supplier Code", "Optional. Supplier's product code"),
    ("K", "Customer Name", "Optional. Must match existing customer"),
    ("L", "Customer Code", "Optional. Customer's product code"),
    ("M", "HS Code", "Optional. Harmonized System Code"),
    ("N", "Country of Origin", "Optional. Country code (e.g., CN, US)"),
    ("O", "Brand", "Optional. Product brand"),
    ("P", "Model Number", "Optional. Model/version number"),
    ("Q", "MOQ", "Optional. Minimum Order Quantity"),
    ("R", "MOQ Unit", "Optional. Unit for MOQ (pcs, cartons, etc.)"),
    ("S", "Lead Time (Days)", "Optional. Production lead time in days"),
    ("T", "Certifications", "Optional. CE, FDA, RoHS, etc."),
    ("U", "Net Weight (kg)", "Optional. Product net weight"),
    ("V", "Gross Weight (kg)", "Optional. Product gross weight"),
    ("W", "Length (cm)", "Optional. Product length"),
    ("X", "Width (cm)", "Optional. Product width"),
    ("Y", "Height (cm)", "Optional. Product height"),
    ("Z", "Pcs per Inner Box", "Optional. Pieces per inner box"),
    ("AA", "Inner Box Length (cm)", "Optional"),
    ("AB", "Inner Box Width (cm)", "Optional"),
    ("AC", "Inner Box Height (cm)", "Optional"),
    ("AD", "Inner Box Weight (kg)", "Optional"),
    ("AE", "Pcs per Carton", "Optional. Pieces per master carton"),
    ("AF", "Inner Boxes per Carton", "Optional"),
    ("AG", "Carton Length (cm)", "Optional"),
    ("AH", "Carton Width (cm)", "Optional"),
    ("AI", "Carton Height (cm)", "Optional"),
    ("AJ", "Carton Weight (kg)", "Optional"),
    ("AK", "Carton CBM", "Optional. Cubic meters"),
    ("AL", "Cartons per 20ft", "Optional. Container loading"),
    ("AM", "Cartons per 40ft", "Optional. Container loading"),
    ("AN", "Cartons per 40HQ", "Optional. Container loading"),
    ("AO", "Packing Notes", "Optional. Special packing instructions"),
    ("AP", "Internal Notes", "Optional. Internal notes"),
    ("AQ", "Tags*", "Required. Comma-separated tags (only first will be used)"),
]

# Styling
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
required_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Write headers
for col, header, description in headers:
    cell = ws[f"{col}1"]
    cell.value = header
    cell.fill = required_fill if "*" in header else header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
    
    # Add comment with description
    from openpyxl.comments import Comment
    cell.comment = Comment(description, "System")

# Set column widths
for col, header, _ in headers:
    ws.column_dimensions[col].width = 18

# Add example data row
example_data = {
    "A": "LED Street Light 100W",
    "B": "LED-100W-001",
    "C": "High-efficiency LED street light with 100W power consumption",
    "D": "active",
    "E": 125.50,
    "F": "USD",
    "G": "https://example.com/images/led-light.jpg",
    "H": "(Paste image here or leave blank)",
    "I": "Shenzhen LED Co",
    "J": "SZ-LED-100",
    "K": "ABC Corp",
    "L": "ABC-LIGHT-01",
    "M": "9405.40",
    "N": "CN",
    "O": "BrightLight",
    "P": "BL-100W-V2",
    "Q": 100,
    "R": "pcs",
    "S": 30,
    "T": "CE, RoHS, IP65",
    "U": 2.5,
    "V": 3.2,
    "W": 45,
    "X": 30,
    "Y": 8,
    "Z": 1,
    "AA": 50,
    "AB": 35,
    "AC": 12,
    "AD": 3.5,
    "AE": 10,
    "AF": 10,
    "AG": 52,
    "AH": 37,
    "AI": 42,
    "AJ": 35,
    "AK": 0.081,
    "AL": 320,
    "AM": 680,
    "AN": 780,
    "AO": "Handle with care, fragile",
    "AP": "High margin product, priority item",
    "AQ": "Electronics, Lighting",
}

for col, value in example_data.items():
    cell = ws[f"{col}2"]
    cell.value = value
    cell.border = border
    cell.alignment = Alignment(vertical="center")

# Add instructions sheet
instructions_ws = wb.create_sheet("Instructions")
instructions_ws.column_dimensions['A'].width = 100

instructions = [
    ("PRODUCT IMPORT TEMPLATE - INSTRUCTIONS", "title"),
    ("", "blank"),
    ("HOW TO USE THIS TEMPLATE:", "header"),
    ("1. Fill in product data starting from row 2 (row 1 contains headers)", "text"),
    ("2. Required fields are marked with * (red background)", "text"),
    ("3. Optional fields can be left empty", "text"),
    ("4. Delete the example row (row 2) before importing", "text"),
    ("", "blank"),
    ("PHOTO IMPORT OPTIONS:", "header"),
    ("Option 1: Photo URL (Column G)", "subheader"),
    ("  - Paste a direct URL to the product image", "text"),
    ("  - Example: https://example.com/images/product.jpg", "text"),
    ("  - Supported formats: JPG, PNG, GIF, WEBP", "text"),
    ("  - Max size: 5MB", "text"),
    ("", "blank"),
    ("Option 2: Embedded Image (Column H)", "subheader"),
    ("  - Copy an image and paste it directly into the cell", "text"),
    ("  - The image will be extracted during import", "text"),
    ("  - Supported formats: JPG, PNG, GIF", "text"),
    ("  - If both URL and embedded image exist, embedded takes priority", "text"),
    ("", "blank"),
    ("IMPORTANT NOTES:", "header"),
    ("- Product Name and Tags are REQUIRED fields", "text"),
    ("- SKU will be auto-generated if left empty", "text"),
    ("- Currency defaults to USD if not specified", "text"),
    ("- Status defaults to 'active' if not specified", "text"),
    ("- Supplier/Customer names must match existing records in the system", "text"),
    ("- Tags should be comma-separated (e.g., 'Electronics, Lighting')", "text"),
    ("- Only the first tag will be used (as per system requirement)", "text"),
    ("- Prices are in currency units, not cents (e.g., 10.50 not 1050)", "text"),
    ("- Decimal fields support up to 2-4 decimal places", "text"),
    ("", "blank"),
    ("EXAMPLE DATA:", "header"),
    ("Row 2 in the 'Product Import Template' sheet contains example data.", "text"),
    ("Review it to understand the expected format, then delete it before importing.", "text"),
    ("", "blank"),
    ("AFTER FILLING THE TEMPLATE:", "header"),
    ("1. Save the file as .xlsx format", "text"),
    ("2. Go to Products page in the system", "text"),
    ("3. Click the 'Import' button", "text"),
    ("4. Upload your Excel file", "text"),
    ("5. Review the import results and any warnings/errors", "text"),
]

row = 1
for text, style in instructions:
    cell = instructions_ws[f"A{row}"]
    cell.value = text
    
    if style == "title":
        cell.font = Font(bold=True, size=16, color="4472C4")
    elif style == "header":
        cell.font = Font(bold=True, size=14, color="4472C4")
    elif style == "subheader":
        cell.font = Font(bold=True, size=12, color="70AD47")
    elif style == "text":
        cell.font = Font(size=11)
    
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    row += 1

# Freeze first row
ws.freeze_panes = "A2"

# Save workbook
output_path = "/home/ubuntu/Impex_project_final/storage/app/public/templates/Product_Import_Template.xlsx"

import os
os.makedirs(os.path.dirname(output_path), exist_ok=True)

wb.save(output_path)
print(f"✅ Template created successfully: {output_path}")
print(f"📁 File size: {os.path.getsize(output_path)} bytes")
